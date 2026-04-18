import os
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from enum import Enum
import streamlit as st


# =============================================================================
# SETTINGS YOU MAY CHANGE
# =============================================================================

# Sheet names to ignore when reading the MDL
IGNORED_SHEETS = [
    "GEP - Cover Page",
    "Revision_history",
    "Instructions",
]

# No search/show-all, so cap the component list to keep UI usable
MAX_COMPONENT_OPTIONS_PER_SHEET = 250

# Sidebar rendering limit (avoid UI overload)
SIDEBAR_COMPONENTS_RENDER = 120


# =============================================================================
# COLUMN INDEX DEFINITIONS
# Column positions in the MDL Excel sheet (0-based, leftmost column = 0)
# =============================================================================

class GEPColumns(Enum):
    Col_Hierarchy1  = 0
    Col_Hierarchy2  = 1
    Col_Hierarchy3  = 2
    Col_DocName     = 3
    Col_FilterComp  = 4
    Col_PackType    = 5
    Col_FileType    = 6
    Col_DocType     = 7


# =============================================================================
# COLUMN INDEX DEFINITIONS FOR ENTRY LISTS
# Each parsed MDL entry (directory or file) is stored as a plain list.
# These Enum values are the indices into that list.
# =============================================================================

class EntryList(Enum):
    Col_Kind            = 0   # "File" or "Directory"
    Col_Product         = 1   # Sheet name / product line
    Col_RelPath         = 2   # Relative path string  e.g. "Product/01_Folder/file.pdf"
    Col_FilterComp      = 3   # Filter / component string from MDL
    Col_PackType        = 4   # Package type string from MDL
    Col_DocType         = 5   # Document type string from MDL
    Col_DocName         = 6   # Document / file name from MDL
    Col_SourceSheet     = 7   # Which Excel sheet this came from
    Col_H1Text          = 8   # H1 directory label  e.g. "01_Overview"
    Col_H2Text          = 9   # H2 directory label  e.g. "02_Reports"
    Col_H3Text          = 10  # H3 directory label  e.g. "03_Sub"
    Col_RepoFound       = 11  # True/False - was a matching file found in the repository?
    Col_RepoFilePath    = 12  # Full path to the matched repository file (or None)


# =============================================================================
# COLUMN INDEX DEFINITIONS FOR REPOSITORY FILE LIST
# Each scanned repository file is stored as a plain list.
# =============================================================================

class RepoFileList(Enum):
    Col_FileName    = 0   # Filename only  e.g. "Report.pdf"
    Col_FullPath    = 1   # Full absolute path  e.g. "C:/Repo/Reports/Report.pdf"


# =============================================================================
# COLOR CONSTANTS FOR EXCEL OUTPUT
# Used in PatternFill calls in the summary Excel generation
# =============================================================================

class openpyXlColors(Enum):
    Color_DarkBlue  = "00274A7F"
    Color_Blue      = "005782BF"
    Color_LightBlue = "00C7D9F2"


# =============================================================================
# PARSE MDL EXCEL
# Opens the MDL workbook, loops through all product sheets,
# and builds two plain lists: arrayEntries (all rows) and two sets.
# Returns: arrayEntries, setProducts, setPackageTypes
# =============================================================================

def parse_mdl(mdl_source):
    """
    Read the MDL Excel file and extract all Directory and File rows.
    Returns a tuple: (arrayEntries, setProducts, setPackageTypes)
    - arrayEntries  : list of lists, one per MDL row, indexed by EntryList enum
    - setProducts   : set of product/sheet names found
    - setPackageTypes: set of package type strings found
    """

    arrayEntries   = []
    setProducts    = set()
    setPackageTypes = set()

    # open workbook in read-only + data_only mode (same as Jochen)
    wb = load_workbook(mdl_source, read_only=True, data_only=True)

    for sheet in wb:

        # skip sheets that are not product sheets
        if sheet.title in IGNORED_SHEETS:
            print("Ignoring sheet: " + sheet.title)
            continue

        print("Reading product sheet: " + sheet.title)
        setProducts.add(sheet.title)

        # track current hierarchy labels while walking rows top-to-bottom
        strCurrentH1 = ""
        strCurrentH2 = ""
        strCurrentH3 = ""

        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True)):

            # guard: row must be long enough to reach the FileType column
            if len(row) <= GEPColumns.Col_FileType.value:
                continue

            # read the FileType cell
            cellFileType = row[GEPColumns.Col_FileType.value]

            # skip empty rows and the header row
            if cellFileType is None:
                continue
            strFileType = str(cellFileType).strip()
            if strFileType == "File Type":
                continue

            # read common fields used for both File and Directory rows
            cellDocName  = row[GEPColumns.Col_DocName.value]
            cellFilter   = row[GEPColumns.Col_FilterComp.value]
            cellPackType = row[GEPColumns.Col_PackType.value]

            strDocName  = "" if cellDocName  is None else str(cellDocName).strip()
            strFilter   = "" if cellFilter   is None else str(cellFilter).strip()
            strPackType = "" if cellPackType is None else str(cellPackType).strip()

            # read DocType (column 7) - may not always be present
            strDocType = ""
            if len(row) > GEPColumns.Col_DocType.value:
                cellDocType = row[GEPColumns.Col_DocType.value]
                if cellDocType is not None:
                    strDocType = str(cellDocType).strip()

            # collect package type into the global set
            if strPackType:
                setPackageTypes.add(strPackType)

            # ------------------------------------------------------------------
            # DIRECTORY rows  - update the current H1 / H2 / H3 labels
            # ------------------------------------------------------------------
            if strFileType == "Directory":

                cellH1 = row[GEPColumns.Col_Hierarchy1.value]
                cellH2 = row[GEPColumns.Col_Hierarchy2.value]
                cellH3 = row[GEPColumns.Col_Hierarchy3.value]

                if cellH1 is not None:
                    # H1 directory - reset H2 and H3
                    try:
                        intH1 = int(cellH1)
                        strCurrentH1 = str(intH1).zfill(2) + "_" + strDocName
                        strCurrentH2 = ""
                        strCurrentH3 = ""
                    except ValueError:
                        print("W: Could not parse H1 value as integer in row " + str(idx+1) + " of sheet " + sheet.title)
                        continue

                elif cellH2 is not None:
                    # H2 directory - reset H3
                    try:
                        intH2 = int(cellH2)
                        strCurrentH2 = str(intH2).zfill(2) + "_" + strDocName
                        strCurrentH3 = ""
                    except ValueError:
                        print("W: Could not parse H2 value as integer in row " + str(idx+1) + " of sheet " + sheet.title)
                        continue

                elif cellH3 is not None:
                    # H3 directory
                    try:
                        intH3 = int(cellH3)
                        strCurrentH3 = str(intH3).zfill(2) + "_" + strDocName
                    except ValueError:
                        print("W: Could not parse H3 value as integer in row " + str(idx+1) + " of sheet " + sheet.title)
                        continue

                # build the relative path for this directory entry
                listParts = [sheet.title, strCurrentH1, strCurrentH2, strCurrentH3]
                listParts = [p for p in listParts if p]
                if not listParts:
                    continue
                strRelPath = "/".join(listParts)

                # store directory entry as a plain list indexed by EntryList enum
                arrayEntries.append([
                    "Directory",          # Col_Kind
                    sheet.title,          # Col_Product
                    strRelPath,           # Col_RelPath
                    strFilter,            # Col_FilterComp
                    strPackType,          # Col_PackType
                    strDocType,           # Col_DocType
                    strDocName,           # Col_DocName
                    sheet.title,          # Col_SourceSheet
                    strCurrentH1,         # Col_H1Text
                    strCurrentH2,         # Col_H2Text
                    strCurrentH3,         # Col_H3Text
                    False,                # Col_RepoFound       (directories don't get matched)
                    None,                 # Col_RepoFilePath
                ])

            # ------------------------------------------------------------------
            # FILE rows  - create a file entry under the current hierarchy
            # ------------------------------------------------------------------
            elif strFileType == "File":

                if not strDocName:
                    print("W: File row with empty document name in row " + str(idx+1) + " of sheet " + sheet.title + " - skipping")
                    continue

                # build relative path including the filename itself
                listParts = [sheet.title, strCurrentH1, strCurrentH2, strCurrentH3, strDocName]
                listParts = [p for p in listParts if p]
                strRelPath = "/".join(listParts)

                # store file entry as a plain list indexed by EntryList enum
                arrayEntries.append([
                    "File",               # Col_Kind
                    sheet.title,          # Col_Product
                    strRelPath,           # Col_RelPath
                    strFilter,            # Col_FilterComp
                    strPackType,          # Col_PackType
                    strDocType,           # Col_DocType
                    strDocName,           # Col_DocName
                    sheet.title,          # Col_SourceSheet
                    strCurrentH1,         # Col_H1Text
                    strCurrentH2,         # Col_H2Text
                    strCurrentH3,         # Col_H3Text
                    False,                # Col_RepoFound       - filled in later by match step
                    None,                 # Col_RepoFilePath    - filled in later by match step
                ])

    wb.close()
    print("MDL parsing complete. Total entries: " + str(len(arrayEntries)))
    return arrayEntries, setProducts, setPackageTypes


# =============================================================================
# HELPER: GET COMPONENT LIST FOR A GIVEN PRODUCT SHEET
# Returns sorted list of unique filter/component strings for the given product
# =============================================================================

def components_for_product(arrayEntries, strProduct):
    """
    Collect all unique filter/component strings for the given product sheet.
    Caps the result at MAX_COMPONENT_OPTIONS_PER_SHEET to keep the UI usable.
    """
    setComps = set()
    for arrayRow in arrayEntries:
        if arrayRow[EntryList.Col_Product.value] == strProduct:
            strComp = arrayRow[EntryList.Col_FilterComp.value]
            if strComp:
                setComps.add(strComp)

    listComps = sorted(setComps, key=lambda x: x.lower())

    if len(listComps) > MAX_COMPONENT_OPTIONS_PER_SHEET:
        listComps = listComps[:MAX_COMPONENT_OPTIONS_PER_SHEET]

    return listComps


# =============================================================================
# HELPER: FILENAME NORMALISATION FOR MATCHING
# Strips extension, lowercases, replaces underscores and dashes with spaces
# so that small naming differences don't break matching.
# =============================================================================

def has_extension(strName):
    """Returns True if the filename has a dot-extension (e.g. '.pdf', '.xlsx')."""
    return "." in os.path.basename(strName)


def normalize_full_name(strName):
    """Normalize a full filename (with extension) for comparison."""
    return strName.strip().lower().replace("_", " ").replace("-", " ")


def normalize_base_name(strName):
    """Normalize a filename without extension for comparison."""
    strBase = os.path.splitext(strName)[0]
    return strBase.strip().lower().replace("_", " ").replace("-", " ")


# =============================================================================
# HELPER: FIND A MATCHING FILE IN THE REPOSITORY FOR ONE MDL FILENAME
# Returns the full path string if found, or None if not found.
# =============================================================================

def find_repository_match(strMdlFilename, arrayRepoFiles):
    """
    Search arrayRepoFiles for a file whose name matches strMdlFilename.
    If the MDL filename has an extension, we match on the full name.
    If the MDL filename has no extension, we match on the base name only.
    Matching is case-insensitive and ignores underscores/dashes.
    """
    boolMdlHasExt  = has_extension(strMdlFilename)
    strMdlFullNorm = normalize_full_name(strMdlFilename)
    strMdlBaseNorm = normalize_base_name(strMdlFilename)

    for arrayRepoRow in arrayRepoFiles:
        strRepoName     = arrayRepoRow[RepoFileList.Col_FileName.value]
        strRepoFullNorm = normalize_full_name(strRepoName)
        strRepoBaseNorm = normalize_base_name(strRepoName)

        if boolMdlHasExt:
            # MDL entry has extension - match on full filename
            if strMdlFullNorm == strRepoFullNorm:
                return arrayRepoRow[RepoFileList.Col_FullPath.value]
        else:
            # MDL entry has no extension - match on base name only
            if strMdlBaseNorm == strRepoBaseNorm:
                return arrayRepoRow[RepoFileList.Col_FullPath.value]

    return None


# =============================================================================
# SCAN REPOSITORY
# Walk the repository root directory and collect all files into a list of lists.
# Returns arrayRepoFiles, each row indexed by RepoFileList enum.
# =============================================================================

def scan_repository(strRepoRoot):
    """
    Walk the repository directory tree and return a list of all files found.
    Each entry is a list: [filename, full_path]
    """
    arrayRepoFiles = []

    for strRoot, listDirs, listFiles in os.walk(strRepoRoot):
        for strFilename in listFiles:
            strFullPath = os.path.join(strRoot, strFilename)
            arrayRepoFiles.append([
                strFilename,   # Col_FileName
                strFullPath,   # Col_FullPath
            ])

    print("Repository scan complete. Files found: " + str(len(arrayRepoFiles)))
    return arrayRepoFiles


# =============================================================================
# MATCH MDL FILE ENTRIES WITH REPOSITORY FILES
# Loops through all File entries in arrayEntries and tries to find
# a matching file in arrayRepoFiles. Updates Col_RepoFound and Col_RepoFilePath
# in-place (same pattern as Jochen updating arrayFilerow in place).
# Returns (intFoundCount, intNotFoundCount)
# =============================================================================

def match_mdl_files_with_repository(arrayEntries, arrayRepoFiles):
    """
    For each File entry in arrayEntries, search arrayRepoFiles for a match.
    Updates the entry list in-place:
      Col_RepoFound     -> True or False
      Col_RepoFilePath  -> full path string or None
    Returns a tuple (intFoundCount, intNotFoundCount).
    """
    intFoundCount    = 0
    intNotFoundCount = 0

    for idx, arrayRow in enumerate(arrayEntries):

        # only process File entries, not Directory entries
        if arrayRow[EntryList.Col_Kind.value] != "File":
            continue

        strDocName   = arrayRow[EntryList.Col_DocName.value]
        strFoundPath = find_repository_match(strDocName, arrayRepoFiles)

        if strFoundPath is not None:
            arrayEntries[idx][EntryList.Col_RepoFound.value]    = True
            arrayEntries[idx][EntryList.Col_RepoFilePath.value]  = strFoundPath
            intFoundCount += 1
        else:
            arrayEntries[idx][EntryList.Col_RepoFound.value]    = False
            arrayEntries[idx][EntryList.Col_RepoFilePath.value]  = None
            intNotFoundCount += 1

    print("Matching complete. Found: " + str(intFoundCount) + "  Not found: " + str(intNotFoundCount))
    return intFoundCount, intNotFoundCount


# =============================================================================
# ENTRY FILTER
# Returns True if this entry should be included in the package,
# based on the user's selected package types and components.
# =============================================================================

def entry_matches(arrayRow, setSelectedPackTypes, setSelectedComponents):
    """
    Returns True if the entry passes both filters:
    - Component filter: entry's (product, filtercomp) must be in the selected set
    - PackType filter: entry's packtype must be in the selected set
    An empty selection set means "no filter applied" (= include everything).
    """
    strProduct  = arrayRow[EntryList.Col_Product.value]
    strFilter   = arrayRow[EntryList.Col_FilterComp.value]
    strPackType = arrayRow[EntryList.Col_PackType.value]

    # component filter (product + filter/component pair)
    if setSelectedComponents and (strProduct, strFilter) not in setSelectedComponents:
        return False

    # package type filter
    if setSelectedPackTypes and strPackType not in setSelectedPackTypes:
        return False

    return True


# =============================================================================
# SAFE PATH JOIN
# Prevents path traversal attacks by verifying the result stays inside base.
# =============================================================================

def safe_join(strBase, *listParts):
    """
    Join strBase with listParts and verify the result is still inside strBase.
    Raises ValueError if the path would escape outside strBase.
    """
    strBaseAbs      = os.path.abspath(strBase)
    strCandidate    = os.path.abspath(os.path.join(strBaseAbs, *listParts))
    if os.path.commonpath([strBaseAbs, strCandidate]) != strBaseAbs:
        raise ValueError("E: Unsafe output path detected - path would escape output folder: " + strCandidate)
    return strCandidate


# =============================================================================
# BUILD PACKAGE
# Loops through all matched File entries, applies filters, and copies
# each file into the output folder structure.
# Returns (intCopied, listMissing, arrayCopiedEntries)
# =============================================================================

def build_package(arrayEntries, strOutputRoot, strProjectName, setSelectedPackTypes, setSelectedComponents):
    """
    Copy files that pass the filter into the output folder.
    - arrayEntries          : full parsed MDL entry list (from parse_mdl)
    - strOutputRoot         : base output directory
    - strProjectName        : project subfolder name inside output root
    - setSelectedPackTypes  : set of package type strings to include (empty = all)
    - setSelectedComponents : set of (product, component) tuples to include
    Returns (intCopied, listMissing, arrayCopiedEntries)
    """
    intCopied          = 0
    listMissing        = []
    arrayCopiedEntries = []

    for arrayRow in arrayEntries:

        # only process File entries
        if arrayRow[EntryList.Col_Kind.value] != "File":
            continue

        # apply component and package type filters
        if not entry_matches(arrayRow, setSelectedPackTypes, setSelectedComponents):
            continue

        # check if this file was matched to a repository file
        boolRepoFound   = arrayRow[EntryList.Col_RepoFound.value]
        strRepoFilePath = arrayRow[EntryList.Col_RepoFilePath.value]

        if not boolRepoFound or strRepoFilePath is None:
            listMissing.append(arrayRow[EntryList.Col_DocName.value])
            continue

        # build destination folder from hierarchy labels
        strH1 = arrayRow[EntryList.Col_H1Text.value]
        strH2 = arrayRow[EntryList.Col_H2Text.value]
        strH3 = arrayRow[EntryList.Col_H3Text.value]

        listFolderParts = [p for p in [strH1, strH2, strH3] if p]

        strDestinationFolder = safe_join(strOutputRoot, strProjectName, *listFolderParts)
        strDestinationFile   = safe_join(strDestinationFolder, os.path.basename(strRepoFilePath))

        # create folder and copy file
        try:
            os.makedirs(strDestinationFolder, exist_ok=True)
            st.write("Copy from: " + strRepoFilePath)
            st.write("Copy to:   " + strDestinationFile)
            shutil.copy2(strRepoFilePath, strDestinationFile)
            intCopied += 1
            arrayCopiedEntries.append(arrayRow)
        except Exception as err:
            st.error("E: Copy failed for \"" + arrayRow[EntryList.Col_DocName.value] + "\": " + str(err))

    print("Build complete. Files copied: " + str(intCopied) + "  Missing: " + str(len(listMissing)))
    return intCopied, listMissing, arrayCopiedEntries


# =============================================================================
# EXPORT SUMMARY EXCEL
# Creates a two-sheet summary workbook:
#   Sheet 1 "Copied Files Summary"  - flat list with [Info] header
#   Sheet 2 "GEP File List"         - grouped by H1/H2/H3 with colors + hyperlinks
# =============================================================================

def export_generated_summary_excel(strOutputRoot, strProjectName, arrayCopiedEntries):
    """
    Generate a two-sheet Excel summary of all copied files.
    Sheet 1: flat list with project info header rows
    Sheet 2: hierarchical list with colored H1/H2/H3 headers and hyperlinks
    Returns the path to the created file, or None if no entries were copied.
    """
    if not arrayCopiedEntries:
        return None

    strProjectFolder = safe_join(strOutputRoot, strProjectName)
    os.makedirs(strProjectFolder, exist_ok=True)

    strSummaryPath = safe_join(strProjectFolder, strProjectName + "_Copied_Files_Summary.xlsx")

    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: "Copied Files Summary" - flat list
    # ------------------------------------------------------------------
    selectionws       = wb.active
    selectionws.title = "Copied Files Summary"

    selectionws.append({"A": "[Info]", "C": "Created for Project:", "D": strProjectName})
    selectionws.append({"A": "[Info]", "C": "Total files copied:",  "D": len(arrayCopiedEntries)})
    selectionws.append({"A": "------------------", "B": "------------------", "C": "------------------"})

    # ------------------------------------------------------------------
    # Sheet 2: "GEP File List" - hierarchical with colors and hyperlinks
    # ------------------------------------------------------------------
    wb.create_sheet("GEP File List")
    summaryws = wb["GEP File List"]

    # header row for sheet 2
    summaryws.append({
        "A": "Structure, Path",
        "B": "Document Name",
        "C": "Document Type",
        "D": "Package Type",
        "E": "Filter/Configuration",
        "F": "relative Link to Document",
    })
    # bold the header row
    for intCol in range(1, 7):
        summaryws.cell(row=summaryws.max_row, column=intCol).font = Font(bold=True)

    # ------------------------------------------------------------------
    # Sort copied entries by product, then H1/H2/H3, then filename
    # (same grouping logic as Jochen's setH1/setH2/setH3 sorted loops)
    # ------------------------------------------------------------------
    arrayCopiedSorted = sorted(
        [r for r in arrayCopiedEntries if r[EntryList.Col_Kind.value] == "File"],
        key=lambda r: (
            r[EntryList.Col_Product.value].lower(),
            r[EntryList.Col_H1Text.value].lower() if r[EntryList.Col_H1Text.value] else "",
            r[EntryList.Col_H2Text.value].lower() if r[EntryList.Col_H2Text.value] else "",
            r[EntryList.Col_H3Text.value].lower() if r[EntryList.Col_H3Text.value] else "",
            r[EntryList.Col_DocName.value].lower(),
        ),
    )

    # sets to track which hierarchy headers have already been written
    setWrittenH1 = set()
    setWrittenH2 = set()
    setWrittenH3 = set()

    for arrayRow in arrayCopiedSorted:

        strH1 = arrayRow[EntryList.Col_H1Text.value]
        strH2 = arrayRow[EntryList.Col_H2Text.value]
        strH3 = arrayRow[EntryList.Col_H3Text.value]

        # H1 header row - dark blue background
        if strH1:
            keyH1 = strH1
            if keyH1 not in setWrittenH1:
                selectionws.append({"A": "[Dir]", "C": strH1})
                summaryws.append({"A": strH1})
                for intCol in range(1, 7):
                    thiscell = summaryws.cell(row=summaryws.max_row, column=intCol)
                    thiscell.font = Font(bold=True, color="FFFFFFFF")
                    thiscell.fill = PatternFill(
                        start_color=openpyXlColors.Color_DarkBlue.value,
                        end_color=openpyXlColors.Color_DarkBlue.value,
                        fill_type="solid",
                    )
                setWrittenH1.add(keyH1)

        # H2 header row - medium blue background
        if strH2:
            keyH2 = (strH1, strH2)
            if keyH2 not in setWrittenH2:
                selectionws.append({"A": "[Dir]", "C": strH2})
                summaryws.append({"A": "  " + strH2})
                for intCol in range(1, 7):
                    thiscell = summaryws.cell(row=summaryws.max_row, column=intCol)
                    thiscell.font = Font(bold=True, color="FFFFFFFF")
                    thiscell.fill = PatternFill(
                        start_color=openpyXlColors.Color_Blue.value,
                        end_color=openpyXlColors.Color_Blue.value,
                        fill_type="solid",
                    )
                setWrittenH2.add(keyH2)

        # H3 header row - light blue background
        if strH3:
            keyH3 = (strH1, strH2, strH3)
            if keyH3 not in setWrittenH3:
                selectionws.append({"A": "[Dir]", "C": strH3})
                summaryws.append({"A": "    " + strH3})
                for intCol in range(1, 7):
                    thiscell = summaryws.cell(row=summaryws.max_row, column=intCol)
                    thiscell.font = Font(bold=True, color="00000000")
                    thiscell.fill = PatternFill(
                        start_color=openpyXlColors.Color_LightBlue.value,
                        end_color=openpyXlColors.Color_LightBlue.value,
                        fill_type="solid",
                    )
                setWrittenH3.add(keyH3)

        # file row - append to both sheets
        strRelLink = arrayRow[EntryList.Col_RelPath.value].replace("\\", "/")

        selectionws.append({
            "A": "[File]",
            "C": arrayRow[EntryList.Col_DocName.value],
            "D": arrayRow[EntryList.Col_DocType.value],
            "E": arrayRow[EntryList.Col_PackType.value],
            "F": strRelLink,
        })

        summaryws.append({
            "A": arrayRow[EntryList.Col_Product.value],
            "B": arrayRow[EntryList.Col_DocName.value],
            "C": arrayRow[EntryList.Col_DocType.value],
            "D": arrayRow[EntryList.Col_PackType.value],
            "E": arrayRow[EntryList.Col_FilterComp.value],
            "F": '=HYPERLINK("./' + strRelLink + '")',
        })

        # wrap text on doc name cell and color the hyperlink cell
        summaryws.cell(row=summaryws.max_row, column=2).alignment = Alignment(wrap_text=True)
        summaryws.cell(row=summaryws.max_row, column=6).font = Font(
            bold=False,
            color=openpyXlColors.Color_DarkBlue.value,
        )

    # ------------------------------------------------------------------
    # Auto-fit column widths on both sheets (cap at 80 to avoid extremes)
    # (same dims pattern as Jochen, plus our width cap)
    # ------------------------------------------------------------------
    for ws in [selectionws, summaryws]:
        dims = {}
        for wsrow in ws.rows:
            for cell in wsrow:
                if cell.value:
                    dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
        for col, value in dims.items():
            ws.column_dimensions[col].width = min(value + 2, 80)

    wb.save(strSummaryPath)
    print("Summary Excel saved: " + strSummaryPath)
    return strSummaryPath


# =============================================================================
# STREAMLIT SESSION STATE HELPERS
# =============================================================================

def init_state():
    """Initialize session state keys if they don't exist yet."""
    if "selected_packtypes" not in st.session_state:
        st.session_state.selected_packtypes = set()
    if "selected_components" not in st.session_state:
        st.session_state.selected_components = set()


def clear_all():
    """Clear all user selections from session state."""
    st.session_state.selected_packtypes  = set()
    st.session_state.selected_components = set()


# =============================================================================
# MAIN STREAMLIT APP
# =============================================================================

def main():
    st.set_page_config(page_title="GEP Package Automation Tool", layout="wide")
    init_state()

    st.title("GEP Package Automation Tool")

    # ---- File upload ----
    uploaded_mdl = st.file_uploader("Upload MDL Excel file", type=["xlsx"])

    if uploaded_mdl is None:
        st.info("Please upload an MDL Excel file to proceed.")
        return

    # parse the uploaded MDL - returns plain lists, not dataclasses
    arrayEntries, setProducts, setPackageTypes = parse_mdl(uploaded_mdl)

    # ---- Sidebar: Review panel ----
    with st.sidebar:
        st.header("Review")

        intPtCount = len(st.session_state.selected_packtypes)
        with st.expander("Package type (" + str(intPtCount) + ")", expanded=False):
            if intPtCount:
                for strPt in sorted(st.session_state.selected_packtypes):
                    st.write("- " + strPt)
            else:
                st.caption("None selected (means ALL package types at build time).")

        listSelectedSorted = sorted(st.session_state.selected_components, key=lambda x: (x[0].lower(), x[1].lower()))
        intCompCount = len(listSelectedSorted)

        with st.expander("Components selected (" + str(intCompCount) + ")", expanded=False):
            if not listSelectedSorted:
                st.caption("No components selected yet.")
            else:
                st.caption("Click x to remove:")
                listToShow = listSelectedSorted[:SIDEBAR_COMPONENTS_RENDER]

                for strProduct, strComp in listToShow:
                    row = st.columns([12, 1])
                    with row[0]:
                        st.write(strProduct + " - " + strComp)
                    with row[1]:
                        if st.button("x", key="rm_" + strProduct + "_" + strComp, help="Remove"):
                            st.session_state.selected_components.discard((strProduct, strComp))
                            st.rerun()

                if intCompCount > SIDEBAR_COMPONENTS_RENDER:
                    st.caption("Showing first " + str(SIDEBAR_COMPONENTS_RENDER) + ".")

        st.markdown("---")
        if st.button("Clear all", use_container_width=True):
            clear_all()
            st.rerun()

    # ---- Main content area ----
    top    = st.container(border=True)
    mid    = st.container(border=True)
    bottom = st.container(border=True)

    # ---- Package Type selection ----
    with top:
        st.subheader("Package Type selection")

        listSelectedPacktypes = st.multiselect(
            "Package Type (Preliminary, Permitting or Operational)",
            options=sorted(setPackageTypes),
            default=sorted(st.session_state.selected_packtypes),
            help="Independent selection. Component options will NOT change based on this.",
        )
        st.session_state.selected_packtypes = set(listSelectedPacktypes)

    # ---- Component selection ----
    with mid:
        st.subheader("Add components")

        col1, col2 = st.columns([1, 2])

        with col1:
            strPickProduct = st.selectbox("Product Type", options=sorted(setProducts))

        with col2:
            listOptions = components_for_product(arrayEntries, strPickProduct)

            if not listOptions:
                st.warning("No components found for this sheet.")
            else:
                if len(listOptions) >= MAX_COMPONENT_OPTIONS_PER_SHEET:
                    st.caption("Showing first " + str(MAX_COMPONENT_OPTIONS_PER_SHEET) + " options (alphabetical).")

                with st.form("add_components_form", clear_on_submit=True):
                    listTempComponents = st.multiselect("Components in " + strPickProduct, options=listOptions)
                    boolSubmitted = st.form_submit_button("Add selected components")

                    if boolSubmitted:
                        for strComp in listTempComponents:
                            st.session_state.selected_components.add((strPickProduct, strComp))
                        st.success("Added " + str(len(listTempComponents)) + " component(s).")
                        st.rerun()

    # ---- Build package ----
    with bottom:
        st.subheader("Build package")

        strRepoRoot    = st.text_input("Repository root",                  value="", placeholder="Path to your repository root folder")
        strOutputRoot  = st.text_input("Output folder (package destination)", value="", placeholder=r"Example: C:\Project\Package_Output")
        strProjectName = st.text_input("Project name",                     value="", placeholder="Example: MyGEPProject")

        if st.button("Build package", type="primary"):

            if not st.session_state.selected_components:
                st.warning("Please select at least one component to build the package.")
                return
            if not strRepoRoot or not os.path.isdir(strRepoRoot):
                st.warning("Please enter a valid repository root path.")
                return
            if not strOutputRoot:
                st.warning("Please enter a valid output folder path.")
                return
            if not strProjectName.strip():
                st.warning("Please enter a valid project name.")
                return

            os.makedirs(strOutputRoot, exist_ok=True)

            with st.spinner("Building package..."):
                try:
                    # step 1: scan the repository for all available files
                    arrayRepoFiles = scan_repository(strRepoRoot)
                    intScannedCount = len(arrayRepoFiles)

                    # step 2: match each MDL file entry to a repository file
                    intMatchedCount, intNotFoundCount = match_mdl_files_with_repository(arrayEntries, arrayRepoFiles)

                    # step 3: copy matching files into the output folder
                    intCopied, listMissing, arrayCopiedEntries = build_package(
                        arrayEntries         = arrayEntries,
                        strOutputRoot        = strOutputRoot,
                        strProjectName       = strProjectName.strip(),
                        setSelectedPackTypes = st.session_state.selected_packtypes,
                        setSelectedComponents= st.session_state.selected_components,
                    )

                    # step 4: create the summary Excel
                    strSummaryExcelPath = export_generated_summary_excel(
                        strOutputRoot        = strOutputRoot,
                        strProjectName       = strProjectName.strip(),
                        arrayCopiedEntries   = arrayCopiedEntries,
                    )

                except ValueError as err:
                    st.error(str(err))
                    return

            st.success(
                "Package build complete! "
                + str(intCopied) + " file(s) copied, "
                + str(len(listMissing)) + " missing, "
                + str(intNotFoundCount) + " MDL file(s) not matched in repository, "
                + "out of " + str(intScannedCount) + " scanned."
            )
            st.info("Package location: " + os.path.abspath(strOutputRoot))

            if strSummaryExcelPath:
                st.success("Summary Excel created at: " + strSummaryExcelPath)
            else:
                st.info("No files were copied, so no summary Excel was created.")

            if listMissing:
                st.warning(str(len(listMissing)) + " MDL files were not found in the repository.")
                with st.expander("See missing files list"):
                    for strMissing in listMissing[:300]:
                        st.write(strMissing)
                    if len(listMissing) > 300:
                        st.caption("Showing first 300 of " + str(len(listMissing)) + " missing files.")


if __name__ == "__main__":
    main()
