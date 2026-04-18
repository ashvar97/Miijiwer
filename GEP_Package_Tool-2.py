import os
import sys
import shutil
import logging
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from enum import Enum
import streamlit as st

try:
    from thefuzz import fuzz
    FUZZY_AVAILABLE = True
except ImportError:
    FUZZY_AVAILABLE = False


# =============================================================================
# SETTINGS YOU MAY CHANGE
# =============================================================================

# Sheet names to ignore when reading the MDL
IGNORED_SHEETS = [
    "GEP - Cover Page",
    "Revision_history",
    "Instructions",
]

# Valid file extensions recognised in MDL filenames (same list as Jochen)
LIST_VALID_EXTENSIONS = ["pdf", "jpg", "docx", "xlsx", "dwg", "step", "stp", "txt", "dxf"]

# Fuzzy match score threshold - if best fuzzy score >= this, log a suggestion (Jochen: 80)
INT_FUZZY_THRESHOLD = 80

# Path length warning threshold (Windows MAX_PATH risk, Jochen: 250)
INT_MAX_PATH_WARNING = 250

# No search/show-all, so cap the component list to keep UI usable
MAX_COMPONENT_OPTIONS_PER_SHEET = 250

# Sidebar rendering limit (avoid UI overload)
SIDEBAR_COMPONENTS_RENDER = 120


# =============================================================================
# LOGGING SETUP
# Creates a timestamped .log file alongside the script AND prints to terminal.
# Same dual-handler pattern as Jochen.
# =============================================================================

current_date = datetime.now()
strLoggingFile = str(
    os.path.basename(__file__) + "_log_" + current_date.strftime("%Y-%m-%d_%H.%M.%S") + ".log"
)

logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

stdout_handler = logging.StreamHandler(sys.stdout)
stdout_handler.setLevel(logging.DEBUG)
stdout_handler.setFormatter(formatter)

file_handler = logging.FileHandler(strLoggingFile)
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(formatter)

logger.addHandler(file_handler)
logger.addHandler(stdout_handler)

logger.info("*********************************************************************")
logger.info("I001 Starting GEP Package Automation Tool")
logger.info("I002 Logging to file: \"" + strLoggingFile + "\"")
logger.info("*********************************************************************")


# =============================================================================
# COLUMN INDEX DEFINITIONS
# Column positions in the MDL Excel sheet (0-based, leftmost column = 0)
# =============================================================================

class GEPColumns(Enum):
    Col_Hierarchy1  = 0
    Col_Hierarchy2  = 1
    Col_Hierarchy3  = 2
    Col_DocName     = 3
    Col_FilterComp  = 4
    Col_PackType    = 5
    Col_FileType    = 6
    Col_DocType     = 7


# =============================================================================
# COLUMN INDEX DEFINITIONS FOR ENTRY LISTS
# Each parsed MDL entry (directory or file) is stored as a plain list.
# These Enum values are the indices into that list.
# =============================================================================

class EntryList(Enum):
    Col_Kind            = 0   # "File" or "Directory"
    Col_Product         = 1   # Sheet name / product line
    Col_RelPath         = 2   # Relative path string  e.g. "Product/01_Folder/file.pdf"
    Col_FilterComp      = 3   # Filter / component string from MDL
    Col_PackType        = 4   # Package type string from MDL
    Col_DocType         = 5   # Document type string from MDL
    Col_DocName         = 6   # Document / file name from MDL
    Col_SourceSheet     = 7   # Which Excel sheet this came from
    Col_H1Text          = 8   # H1 directory label  e.g. "01_Overview"
    Col_H2Text          = 9   # H2 directory label  e.g. "02_Reports"
    Col_H3Text          = 10  # H3 directory label  e.g. "03_Sub"
    Col_RepoFound       = 11  # True/False - was a matching file found in the repository?
    Col_RepoFilePath    = 12  # Full path to the matched repository file (or None)


# =============================================================================
# COLUMN INDEX DEFINITIONS FOR REPOSITORY FILE LIST
# Each scanned repository file is stored as a plain list.
# =============================================================================

class RepoFileList(Enum):
    Col_FileName    = 0   # Filename only  e.g. "Report.pdf"
    Col_FullPath    = 1   # Full absolute path  e.g. "C:/Repo/Reports/Report.pdf"
    Col_UsedinMDL   = 2   # True/False - has this file been matched to an MDL entry?


# =============================================================================
# COLOR CONSTANTS FOR EXCEL OUTPUT
# Used in PatternFill calls in the summary Excel generation
# =============================================================================

class openpyXlColors(Enum):
    Color_DarkBlue  = "00274A7F"
    Color_Blue      = "005782BF"
    Color_LightBlue = "00C7D9F2"


# =============================================================================
# PARSE MDL EXCEL
# Opens the MDL workbook, loops through all product sheets,
# and builds a list of entry rows plus two sets (products, package types).
# Returns: (arrayEntries, setProducts, setPackageTypes)
# =============================================================================

def parse_mdl(mdl_source):
    """
    Read the MDL Excel file and extract all Directory and File rows.
    Returns a tuple: (arrayEntries, setProducts, setPackageTypes)
    - arrayEntries   : list of lists, one per MDL row, indexed by EntryList enum
    - setProducts    : set of product/sheet names found
    - setPackageTypes: set of package type strings found
    """

    logger.info("*********************************************************************")
    logger.info("I010 Read in MDL Excel Spreadsheet")
    logger.info("*********************************************************************")

    arrayEntries      = []
    setProducts       = set()
    setPackageTypes   = set()
    intMDLDirectories = 0
    intMDLFiles       = 0

    # open workbook in read-only + data_only mode (same as Jochen)
    wb = load_workbook(mdl_source, read_only=True, data_only=True)

    for sheet in wb:

        # skip sheets that are not product sheets
        if sheet.title in IGNORED_SHEETS:
            logger.info("I011 Ignoring sheet: \"" + sheet.title + "\"")
            continue

        logger.info("I012 Reading product sheet: \"" + sheet.title + "\"")
        setProducts.add(sheet.title)

        # track current hierarchy labels while walking rows top-to-bottom
        strCurrentH1 = ""
        strCurrentH2 = ""
        strCurrentH3 = ""

        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True)):

            # guard: row must be long enough to reach the FileType column
            if len(row) <= GEPColumns.Col_FileType.value:
                continue

            # ------------------------------------------------------------------
            # Non-printable character check on every cell in this row (Jochen W021)
            # ------------------------------------------------------------------
            for idy, cellval in enumerate(row):
                strCellCheck = str(cellval)
                if not strCellCheck.isprintable():
                    logger.warning(
                        "W013 Non-printable characters in cell value. "
                        "Sheet: \"" + sheet.title + "\" "
                        "Row: " + str(idx + 1) + " Col: " + str(idy + 1) + " "
                        "Value: " + strCellCheck
                    )

            # read the FileType cell
            cellFileType = row[GEPColumns.Col_FileType.value]

            # skip empty rows and the header row
            if cellFileType is None:
                continue
            strFileType = str(cellFileType).strip()
            if strFileType == "File Type":
                continue

            # read common fields used for both File and Directory rows
            cellDocName  = row[GEPColumns.Col_DocName.value]
            cellFilter   = row[GEPColumns.Col_FilterComp.value]
            cellPackType = row[GEPColumns.Col_PackType.value]

            strDocName  = "" if cellDocName  is None else str(cellDocName).strip()
            strFilter   = "" if cellFilter   is None else str(cellFilter).strip()
            strPackType = "" if cellPackType is None else str(cellPackType).strip()

            # read DocType (column 7) - may not always be present
            strDocType = ""
            if len(row) > GEPColumns.Col_DocType.value:
                cellDocType = row[GEPColumns.Col_DocType.value]
                if cellDocType is not None:
                    strDocType = str(cellDocType).strip()

            # collect package type into the global set
            if strPackType:
                setPackageTypes.add(strPackType)

            # ------------------------------------------------------------------
            # DIRECTORY rows - update the current H1 / H2 / H3 labels
            # ------------------------------------------------------------------
            if strFileType == "Directory":

                cellH1 = row[GEPColumns.Col_Hierarchy1.value]
                cellH2 = row[GEPColumns.Col_Hierarchy2.value]
                cellH3 = row[GEPColumns.Col_Hierarchy3.value]

                boolCanUseRow = True

                if cellH1 is not None:
                    # H1 directory - reset H2 and H3
                    try:
                        intH1 = int(cellH1)
                        strCurrentH1 = str("{:02d}".format(intH1)) + "_" + strDocName
                        strCurrentH1 = strCurrentH1.strip().replace('"', "").replace("|", "")
                        strCurrentH2 = ""
                        strCurrentH3 = ""
                        logger.debug(
                            "D014 New H1 directory: \"" + strCurrentH1 + "\" "
                            "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1)
                        )
                    except ValueError:
                        boolCanUseRow = False
                        logger.error(
                            "E015 Non-integer value in Hierarchy H1. "
                            "Sheet: \"" + sheet.title + "\" "
                            "Row: " + str(idx + 1) + " Value: \"" + str(cellH1) + "\""
                        )

                elif cellH2 is not None:
                    # H2 directory - reset H3
                    try:
                        intH2 = int(cellH2)
                        strCurrentH2 = str("{:02d}".format(intH2)) + "_" + strDocName
                        strCurrentH2 = strCurrentH2.strip().replace('"', "").replace("|", "")
                        strCurrentH3 = ""
                        logger.debug(
                            "D016 New H2 directory: \"" + strCurrentH2 + "\" "
                            "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1)
                        )
                    except ValueError:
                        boolCanUseRow = False
                        logger.error(
                            "E017 Non-integer value in Hierarchy H2. "
                            "Sheet: \"" + sheet.title + "\" "
                            "Row: " + str(idx + 1) + " Value: \"" + str(cellH2) + "\""
                        )

                elif cellH3 is not None:
                    # H3 directory
                    try:
                        intH3 = int(cellH3)
                        strCurrentH3 = str("{:02d}".format(intH3)) + "_" + strDocName
                        strCurrentH3 = strCurrentH3.strip().replace('"', "").replace("|", "")
                        logger.debug(
                            "D018 New H3 directory: \"" + strCurrentH3 + "\" "
                            "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1)
                        )
                    except ValueError:
                        boolCanUseRow = False
                        logger.error(
                            "E019 Non-integer value in Hierarchy H3. "
                            "Sheet: \"" + sheet.title + "\" "
                            "Row: " + str(idx + 1) + " Value: \"" + str(cellH3) + "\""
                        )

                if not boolCanUseRow:
                    logger.error(
                        "E020 Discarded Directory row due to previous error. "
                        "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1)
                    )
                    continue

                # build the relative path for this directory entry
                listParts = [sheet.title, strCurrentH1, strCurrentH2, strCurrentH3]
                listParts = [p for p in listParts if p]
                if not listParts:
                    continue
                strRelPath = "/".join(listParts)

                # store directory entry as a plain list indexed by EntryList enum
                arrayEntries.append([
                    "Directory",     # Col_Kind
                    sheet.title,     # Col_Product
                    strRelPath,      # Col_RelPath
                    strFilter,       # Col_FilterComp
                    strPackType,     # Col_PackType
                    strDocType,      # Col_DocType
                    strDocName,      # Col_DocName
                    sheet.title,     # Col_SourceSheet
                    strCurrentH1,    # Col_H1Text
                    strCurrentH2,    # Col_H2Text
                    strCurrentH3,    # Col_H3Text
                    False,           # Col_RepoFound    (directories are not matched)
                    None,            # Col_RepoFilePath
                ])
                intMDLDirectories += 1

            # ------------------------------------------------------------------
            # FILE rows - create a file entry under the current hierarchy
            # ------------------------------------------------------------------
            elif strFileType == "File":

                boolCanUseRow = True

                # validate: document name must not be empty (Jochen E031)
                if not strDocName:
                    boolCanUseRow = False
                    logger.error(
                        "E021 File row has empty document name. "
                        "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1)
                    )

                # validate: FilterComp must not be empty (Jochen E033)
                if cellFilter is None or strFilter == "":
                    boolCanUseRow = False
                    logger.error(
                        "E022 Missing Filter/Component setting for File row. "
                        "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1) + " "
                        "Hierarchy: " + strCurrentH1 + "." + strCurrentH2 + "." + strCurrentH3
                    )

                # validate: PackType must not be None, empty, or "NA" (Jochen E034/E035)
                if cellPackType is None or strPackType == "" or strPackType == "NA":
                    boolCanUseRow = False
                    logger.error(
                        "E023 Invalid or missing Package Type for File row. "
                        "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1) + " "
                        "Value: \"" + str(strPackType) + "\""
                    )

                # info: multiple dots in filename may indicate extension confusion (Jochen I032)
                if strDocName:
                    intDots = strDocName.count(".")
                    if intDots > 1:
                        logger.info(
                            "I024 Found " + str(intDots) + " dots in document name - "
                            "check for extension conflicts. "
                            "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1) + " "
                            "Name: \"" + strDocName + "\""
                        )

                if not boolCanUseRow:
                    logger.error(
                        "E025 Discarded File row due to previous error. "
                        "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1) + " "
                        "Name: \"" + strDocName + "\""
                    )
                    continue

                # build relative path including the filename itself
                listParts = [sheet.title, strCurrentH1, strCurrentH2, strCurrentH3, strDocName]
                listParts = [p for p in listParts if p]
                strRelPath = "/".join(listParts)

                # store file entry as a plain list indexed by EntryList enum
                arrayEntries.append([
                    "File",          # Col_Kind
                    sheet.title,     # Col_Product
                    strRelPath,      # Col_RelPath
                    strFilter,       # Col_FilterComp
                    strPackType,     # Col_PackType
                    strDocType,      # Col_DocType
                    strDocName,      # Col_DocName
                    sheet.title,     # Col_SourceSheet
                    strCurrentH1,    # Col_H1Text
                    strCurrentH2,    # Col_H2Text
                    strCurrentH3,    # Col_H3Text
                    False,           # Col_RepoFound    - filled in later by match step
                    None,            # Col_RepoFilePath - filled in later by match step
                ])
                intMDLFiles += 1

            # ------------------------------------------------------------------
            # Unknown / faulty row type - warn if row is not empty (Jochen W037)
            # ------------------------------------------------------------------
            else:
                boolRowEmpty = all(item is None for item in row)
                if not boolRowEmpty:
                    logger.warning(
                        "W026 Non-empty row has unrecognised File Type entry. "
                        "Sheet: \"" + sheet.title + "\" Row: " + str(idx + 1) + " "
                        "FileType value: \"" + str(strFileType) + "\""
                    )

    wb.close()

    logger.info("*********************************************************************")
    logger.info("I027 MDL Parsing Summary:")
    logger.info("I027 Directories parsed: " + str(intMDLDirectories))
    logger.info("I027 Files parsed:       " + str(intMDLFiles))
    logger.info("I027 Total entries:      " + str(len(arrayEntries)))
    logger.info("*********************************************************************")

    return arrayEntries, setProducts, setPackageTypes


# =============================================================================
# HELPER: GET COMPONENT LIST FOR A GIVEN PRODUCT SHEET
# Returns sorted list of unique filter/component strings for the given product.
# =============================================================================

def components_for_product(arrayEntries, strProduct):
    """
    Collect all unique filter/component strings for the given product sheet.
    Caps the result at MAX_COMPONENT_OPTIONS_PER_SHEET to keep the UI usable.
    """
    setComps = set()
    for arrayRow in arrayEntries:
        if arrayRow[EntryList.Col_Product.value] == strProduct:
            strComp = arrayRow[EntryList.Col_FilterComp.value]
            if strComp:
                setComps.add(strComp)

    listComps = sorted(setComps, key=lambda x: x.lower())

    if len(listComps) > MAX_COMPONENT_OPTIONS_PER_SHEET:
        listComps = listComps[:MAX_COMPONENT_OPTIONS_PER_SHEET]

    return listComps


# =============================================================================
# HELPER: EXTENSION VALIDATION
# Checks whether a filename has a recognised valid extension (Jochen's whitelist).
# Returns a tuple: (strExtStatus, boolHasValidExt)
# =============================================================================

def check_mdl_extension(strFilename):
    """
    Determine whether the MDL filename has a valid file extension.
    Returns a tuple: (strExtStatus, boolHasValidExt)
    - strExtStatus    : "valid_ext", "no_ext", or "invalid_ext"
    - boolHasValidExt : True only if extension is in LIST_VALID_EXTENSIONS and 2-4 chars
    """
    intDots = strFilename.count(".")
    if intDots == 0:
        return "no_ext", False

    strExt = str(strFilename.rsplit(".", 1)[1]).lower()

    if strExt in LIST_VALID_EXTENSIONS and 2 <= len(strExt) <= 4:
        return "valid_ext", True
    elif len(strExt) > 4:
        # very long "extension" - likely part of the filename, not a real extension
        return "no_ext", False
    else:
        return "invalid_ext", False


# =============================================================================
# HELPER: FILENAME NORMALISATION FOR MATCHING
# Strips extension, lowercases, replaces underscores and dashes with spaces.
# =============================================================================

def normalize_full_name(strName):
    """Normalize a full filename (with extension) for comparison."""
    return strName.strip().lower().replace("_", " ").replace("-", " ")


def normalize_base_name(strName):
    """Normalize a filename without its extension for comparison."""
    strBase = os.path.splitext(strName)[0]
    return strBase.strip().lower().replace("_", " ").replace("-", " ")


# =============================================================================
# SCAN REPOSITORY
# Walk the repository root directory and collect all files into a list of lists.
# Also validates path length and file readability (Jochen W050, E051).
# Returns arrayRepoFiles, each row indexed by RepoFileList enum.
# =============================================================================

def scan_repository(strRepoRoot):
    """
    Walk the repository directory tree and return a list of all files found.
    Each entry is a list: [filename, full_path, bool_used_in_mdl]
    Also warns on long paths and logs unreadable files.
    """
    logger.info("*********************************************************************")
    logger.info("I030 Scan Repository Files")
    logger.info("*********************************************************************")

    arrayRepoFiles  = []
    intFilesValid   = 0
    intFilesInvalid = 0
    intMaxPathLen   = 0

    for strRoot, listDirs, listFiles in os.walk(strRepoRoot):
        for strFilename in listFiles:
            strFullPath = os.path.join(strRoot, strFilename)

            # path length warning (Jochen W050)
            intPathLen = len(strFullPath)
            if intPathLen > INT_MAX_PATH_WARNING:
                if intPathLen > intMaxPathLen:
                    intMaxPathLen = intPathLen
                logger.warning(
                    "W031 File path close to or exceeding " + str(INT_MAX_PATH_WARNING) + " characters. "
                    "Length: " + str(intPathLen) + " Path: \"" + strFullPath + "\""
                )

            # file readability check (Jochen E051)
            try:
                with open(strFullPath, "rb") as fp:
                    pass
                intFilesValid += 1
            except Exception as err:
                logger.error(
                    "E032 Cannot open/read repository file: \"" + strFullPath + "\" Error: " + str(err)
                )
                intFilesInvalid += 1
                continue

            arrayRepoFiles.append([
                strFilename,   # Col_FileName
                strFullPath,   # Col_FullPath
                False,         # Col_UsedinMDL - updated later during matching
            ])

    logger.info("*********************************************************************")
    logger.info("I033 Repository Scan Summary:")
    logger.info("I033 Valid files found:   " + str(intFilesValid))
    logger.info("I033 Unreadable files:    " + str(intFilesInvalid))
    logger.info("I033 Max path length:     " + str(intMaxPathLen) + " characters")
    logger.info("*********************************************************************")

    return arrayRepoFiles


# =============================================================================
# MATCH MDL FILE ENTRIES WITH REPOSITORY FILES
# For each MDL File entry, searches the repository for a matching file.
# Updates Col_RepoFound and Col_RepoFilePath in-place on arrayEntries.
# Also updates Col_UsedinMDL in-place on arrayRepoFiles (for reverse check).
# Includes fuzzy match suggestion and multiple-match warning (Jochen style).
# Returns (intFoundCount, intNotFoundCount)
# =============================================================================

def match_mdl_files_with_repository(arrayEntries, arrayRepoFiles):
    """
    For each File entry in arrayEntries, search arrayRepoFiles for a match.
    Updates entries in-place. Also marks matched repo files as used.
    Returns (intFoundCount, intNotFoundCount).
    """
    logger.info("*********************************************************************")
    logger.info("I040 Match MDL Files with Repository Files")
    logger.info("*********************************************************************")

    intFoundCount         = 0
    intNotFoundCount      = 0
    intFuzzyCount         = 0
    intMultipleFoundCount = 0

    for idx, arrayRow in enumerate(arrayEntries):

        # only process File entries, not Directory entries
        if arrayRow[EntryList.Col_Kind.value] != "File":
            continue

        strDocName = arrayRow[EntryList.Col_DocName.value]

        # check whether the MDL filename has a valid extension (Jochen's whitelist)
        strExtStatus, boolHasValidExt = check_mdl_extension(strDocName)

        if strExtStatus == "invalid_ext":
            logger.info(
                "I041 MDL entry \"" + strDocName + "\" has extension not in valid list "
                + str(LIST_VALID_EXTENSIONS) + " - will attempt base-name matching only."
            )

        # normalise MDL filename for comparison
        strMdlFullNorm = normalize_full_name(strDocName)
        strMdlBaseNorm = normalize_base_name(strDocName)

        intFound        = 0
        intBestScore    = 0
        intBestMatchIdx = 0

        for idy, arrayRepoRow in enumerate(arrayRepoFiles):
            strRepoFilename = arrayRepoRow[RepoFileList.Col_FileName.value]
            strRepoFullNorm = normalize_full_name(strRepoFilename)
            strRepoBaseNorm = normalize_base_name(strRepoFilename)

            boolMatched = False

            if boolHasValidExt:
                # MDL entry has valid extension - match on full filename (Jochen best case)
                if strMdlFullNorm == strRepoFullNorm:
                    logger.info(
                        "I042 Best case match (full name): MDL \"" + strDocName + "\" "
                        "-> Repo \"" + strRepoFilename + "\""
                    )
                    boolMatched = True
            else:
                # MDL entry has no valid extension - match on base name (Jochen second best)
                if strMdlBaseNorm == strRepoBaseNorm:
                    logger.info(
                        "I043 Second best match (base name): MDL \"" + strDocName + "\" "
                        "-> Repo \"" + strRepoFilename + "\""
                    )
                    boolMatched = True

            if boolMatched:
                intFound += 1
                # store first match found into the entry
                if intFound == 1:
                    arrayEntries[idx][EntryList.Col_RepoFound.value]   = True
                    arrayEntries[idx][EntryList.Col_RepoFilePath.value] = arrayRepoRow[RepoFileList.Col_FullPath.value]
                    # mark this repo file as used in MDL (for the reverse check)
                    arrayRepoFiles[idy][RepoFileList.Col_UsedinMDL.value] = True

            # fuzzy score tracking for unmatched suggestions (Jochen I065)
            if FUZZY_AVAILABLE:
                intScore = fuzz.ratio(strRepoFilename, strDocName)
                if intScore > intBestScore:
                    intBestScore    = intScore
                    intBestMatchIdx = idy

        # multiple match warning (Jochen W066)
        if intFound > 1:
            logger.warning(
                "W044 MDL entry \"" + strDocName + "\" matched " + str(intFound) + " repository files. "
                "First match will be used. Check for duplicate filenames in repository."
            )
            intMultipleFoundCount += 1

        if intFound >= 1:
            intFoundCount += 1
        else:
            # not found - log warning and fuzzy suggestion if available (Jochen W064 / I065)
            arrayEntries[idx][EntryList.Col_RepoFound.value]   = False
            arrayEntries[idx][EntryList.Col_RepoFilePath.value] = None
            logger.warning(
                "W045 MDL file not found in repository: \"" + strDocName + "\""
            )
            if FUZZY_AVAILABLE and intBestScore >= INT_FUZZY_THRESHOLD:
                strBestName = arrayRepoFiles[intBestMatchIdx][RepoFileList.Col_FileName.value]
                logger.info(
                    "I046 --> Best fuzzy match suggestion (score=" + str(intBestScore) + "%) "
                    "for \"" + strDocName + "\" is: \"" + strBestName + "\""
                )
                intFuzzyCount += 1
            intNotFoundCount += 1

    # ------------------------------------------------------------------
    # Reverse check - flag repo files not referenced by any MDL entry (Jochen W077)
    # ------------------------------------------------------------------
    logger.info("*********************************************************************")
    logger.info("I047 Reverse Check - Repository files not referenced in MDL")
    logger.info("*********************************************************************")

    intUsedInMDL   = 0
    intUnusedInMDL = 0

    for arrayRepoRow in arrayRepoFiles:
        if arrayRepoRow[RepoFileList.Col_UsedinMDL.value] == True:
            intUsedInMDL += 1
        else:
            logger.warning(
                "W048 Repository file not referenced in MDL "
                "(possible typo or outdated file): \""
                + arrayRepoRow[RepoFileList.Col_FullPath.value] + "\""
            )
            intUnusedInMDL += 1

    logger.info("*********************************************************************")
    logger.info("I049 Matching Summary:")
    logger.info("I049 MDL files found in repository:          " + str(intFoundCount))
    logger.info("I049 MDL files NOT found in repository:      " + str(intNotFoundCount))
    logger.info("I049 MDL files with fuzzy match suggestion:  " + str(intFuzzyCount))
    logger.info("I049 MDL files matched multiple times:       " + str(intMultipleFoundCount))
    logger.info("I049 Repository files referenced in MDL:     " + str(intUsedInMDL))
    logger.info("I049 Repository files NOT in MDL (orphaned): " + str(intUnusedInMDL))
    logger.info("*********************************************************************")

    return intFoundCount, intNotFoundCount


# =============================================================================
# ENTRY FILTER
# Returns True if this entry should be included in the package,
# based on the user's selected package types and components.
# =============================================================================

def entry_matches(arrayRow, setSelectedPackTypes, setSelectedComponents):
    """
    Returns True if the entry passes both filters:
    - Component filter: entry's (product, filtercomp) must be in the selected set
    - PackType filter:  entry's packtype must be in the selected set
    An empty selection set means "no filter applied" (include everything).
    """
    strProduct  = arrayRow[EntryList.Col_Product.value]
    strFilter   = arrayRow[EntryList.Col_FilterComp.value]
    strPackType = arrayRow[EntryList.Col_PackType.value]

    # component filter (product + filter/component pair)
    if setSelectedComponents and (strProduct, strFilter) not in setSelectedComponents:
        return False

    # package type filter
    if setSelectedPackTypes and strPackType not in setSelectedPackTypes:
        return False

    return True


# =============================================================================
# SAFE PATH JOIN
# Prevents path traversal attacks by verifying the result stays inside base.
# =============================================================================

def safe_join(strBase, *listParts):
    """
    Join strBase with listParts and verify the result is still inside strBase.
    Raises ValueError if the path would escape outside strBase.
    """
    strBaseAbs   = os.path.abspath(strBase)
    strCandidate = os.path.abspath(os.path.join(strBaseAbs, *listParts))
    if os.path.commonpath([strBaseAbs, strCandidate]) != strBaseAbs:
        raise ValueError(
            "E060 Unsafe output path detected - path would escape output folder: " + strCandidate
        )
    return strCandidate


# =============================================================================
# BUILD PACKAGE
# Loops through all matched File entries, applies filters, and copies
# each file into the output folder structure.
# Returns (intCopied, listMissing, arrayCopiedEntries)
# =============================================================================

def build_package(arrayEntries, strOutputRoot, strProjectName, setSelectedPackTypes, setSelectedComponents):
    """
    Copy files that pass the filter into the output folder.
    Returns (intCopied, listMissing, arrayCopiedEntries).
    """
    logger.info("*********************************************************************")
    logger.info("I061 Assemble Package")
    logger.info("*********************************************************************")

    intCopied          = 0
    listMissing        = []
    arrayCopiedEntries = []

    for arrayRow in arrayEntries:

        # only process File entries
        if arrayRow[EntryList.Col_Kind.value] != "File":
            continue

        # apply component and package type filters
        if not entry_matches(arrayRow, setSelectedPackTypes, setSelectedComponents):
            logger.debug(
                "D062 Skipped by filter: \"" + arrayRow[EntryList.Col_DocName.value] + "\""
            )
            continue

        # check if this file was matched to a repository file
        boolRepoFound   = arrayRow[EntryList.Col_RepoFound.value]
        strRepoFilePath = arrayRow[EntryList.Col_RepoFilePath.value]

        if not boolRepoFound or strRepoFilePath is None:
            listMissing.append(arrayRow[EntryList.Col_DocName.value])
            logger.warning(
                "W063 File selected for package but not found in repository: \""
                + arrayRow[EntryList.Col_DocName.value] + "\""
            )
            continue

        # build destination folder path from hierarchy labels
        strH1 = arrayRow[EntryList.Col_H1Text.value]
        strH2 = arrayRow[EntryList.Col_H2Text.value]
        strH3 = arrayRow[EntryList.Col_H3Text.value]
        listFolderParts = [p for p in [strH1, strH2, strH3] if p]

        strDestinationFolder = safe_join(strOutputRoot, strProjectName, *listFolderParts)
        strDestinationFile   = safe_join(strDestinationFolder, os.path.basename(strRepoFilePath))

        # create folder and copy file (Jochen I087 / E088 pattern)
        try:
            os.makedirs(strDestinationFolder, exist_ok=True)
            logger.info(
                "I064 Copying \"" + strRepoFilePath + "\" "
                "to \"" + strDestinationFile + "\""
            )
            st.write("Copy from: " + strRepoFilePath)
            st.write("Copy to:   " + strDestinationFile)
            shutil.copy2(strRepoFilePath, strDestinationFile)
            intCopied += 1
            arrayCopiedEntries.append(arrayRow)
        except Exception as err:
            logger.error(
                "E065 Cannot copy file \"" + strRepoFilePath + "\" "
                "to \"" + strDestinationFile + "\" Error: " + str(err)
            )
            st.error("E065 Copy failed for \"" + arrayRow[EntryList.Col_DocName.value] + "\": " + str(err))

    logger.info("*********************************************************************")
    logger.info("I066 Package Build Summary:")
    logger.info("I066 Files copied:  " + str(intCopied))
    logger.info("I066 Files missing: " + str(len(listMissing)))
    logger.info("*********************************************************************")

    return intCopied, listMissing, arrayCopiedEntries


# =============================================================================
# EXPORT SUMMARY EXCEL
# Creates a two-sheet summary workbook:
#   Sheet 1 "Copied Files Summary"  - flat list with [Info] header
#   Sheet 2 "GEP File List"         - grouped by H1/H2/H3 with colors + hyperlinks
# =============================================================================

def export_generated_summary_excel(strOutputRoot, strProjectName, arrayCopiedEntries):
    """
    Generate a two-sheet Excel summary of all copied files.
    Returns the path to the created file, or None if no entries were copied.
    """
    logger.info("*********************************************************************")
    logger.info("I070 Export Summary Excel")
    logger.info("*********************************************************************")

    if not arrayCopiedEntries:
        logger.info("I071 No copied entries - summary Excel will not be created.")
        return None

    strProjectFolder = safe_join(strOutputRoot, strProjectName)
    os.makedirs(strProjectFolder, exist_ok=True)

    strSummaryPath = safe_join(strProjectFolder, strProjectName + "_Copied_Files_Summary.xlsx")

    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: "Copied Files Summary" - flat list
    # ------------------------------------------------------------------
    selectionws       = wb.active
    selectionws.title = "Copied Files Summary"

    selectionws.append({"A": "[Info]", "C": "Created for Project:", "D": strProjectName})
    selectionws.append({"A": "[Info]", "C": "Total files copied:",  "D": len(arrayCopiedEntries)})
    selectionws.append({"A": "[Info]", "C": "Created on:",          "D": current_date.strftime("%Y-%m-%d %H:%M:%S")})
    selectionws.append({"A": "------------------", "B": "------------------", "C": "------------------"})

    # ------------------------------------------------------------------
    # Sheet 2: "GEP File List" - hierarchical with colors and hyperlinks
    # ------------------------------------------------------------------
    wb.create_sheet("GEP File List")
    summaryws = wb["GEP File List"]

    # header row for sheet 2
    summaryws.append({
        "A": "Structure, Path",
        "B": "Document Name",
        "C": "Document Type",
        "D": "Package Type",
        "E": "Filter/Configuration",
        "F": "relative Link to Document",
    })
    # bold the header row
    for intCol in range(1, 7):
        summaryws.cell(row=summaryws.max_row, column=intCol).font = Font(bold=True)

    # ------------------------------------------------------------------
    # Sort copied entries by product, then H1/H2/H3, then filename
    # ------------------------------------------------------------------
    arrayCopiedSorted = sorted(
        [r for r in arrayCopiedEntries if r[EntryList.Col_Kind.value] == "File"],
        key=lambda r: (
            r[EntryList.Col_Product.value].lower(),
            r[EntryList.Col_H1Text.value].lower() if r[EntryList.Col_H1Text.value] else "",
            r[EntryList.Col_H2Text.value].lower() if r[EntryList.Col_H2Text.value] else "",
            r[EntryList.Col_H3Text.value].lower() if r[EntryList.Col_H3Text.value] else "",
            r[EntryList.Col_DocName.value].lower(),
        ),
    )

    # sets to track which hierarchy headers have already been written
    setWrittenH1 = set()
    setWrittenH2 = set()
    setWrittenH3 = set()

    for arrayRow in arrayCopiedSorted:

        strH1 = arrayRow[EntryList.Col_H1Text.value]
        strH2 = arrayRow[EntryList.Col_H2Text.value]
        strH3 = arrayRow[EntryList.Col_H3Text.value]

        # H1 header row - dark blue background
        if strH1:
            keyH1 = strH1
            if keyH1 not in setWrittenH1:
                selectionws.append({"A": "[Dir]", "C": strH1})
                summaryws.append({"A": strH1})
                for intCol in range(1, 7):
                    thiscell = summaryws.cell(row=summaryws.max_row, column=intCol)
                    thiscell.font = Font(bold=True, color="FFFFFFFF")
                    thiscell.fill = PatternFill(
                        start_color=openpyXlColors.Color_DarkBlue.value,
                        end_color=openpyXlColors.Color_DarkBlue.value,
                        fill_type="solid",
                    )
                setWrittenH1.add(keyH1)

        # H2 header row - medium blue background
        if strH2:
            keyH2 = (strH1, strH2)
            if keyH2 not in setWrittenH2:
                selectionws.append({"A": "[Dir]", "C": strH2})
                summaryws.append({"A": "  " + strH2})
                for intCol in range(1, 7):
                    thiscell = summaryws.cell(row=summaryws.max_row, column=intCol)
                    thiscell.font = Font(bold=True, color="FFFFFFFF")
                    thiscell.fill = PatternFill(
                        start_color=openpyXlColors.Color_Blue.value,
                        end_color=openpyXlColors.Color_Blue.value,
                        fill_type="solid",
                    )
                setWrittenH2.add(keyH2)

        # H3 header row - light blue background
        if strH3:
            keyH3 = (strH1, strH2, strH3)
            if keyH3 not in setWrittenH3:
                selectionws.append({"A": "[Dir]", "C": strH3})
                summaryws.append({"A": "    " + strH3})
                for intCol in range(1, 7):
                    thiscell = summaryws.cell(row=summaryws.max_row, column=intCol)
                    thiscell.font = Font(bold=True, color="00000000")
                    thiscell.fill = PatternFill(
                        start_color=openpyXlColors.Color_LightBlue.value,
                        end_color=openpyXlColors.Color_LightBlue.value,
                        fill_type="solid",
                    )
                setWrittenH3.add(keyH3)

        # file row - append to both sheets
        strRelLink = arrayRow[EntryList.Col_RelPath.value].replace("\\", "/")

        selectionws.append({
            "A": "[File]",
            "C": arrayRow[EntryList.Col_DocName.value],
            "D": arrayRow[EntryList.Col_DocType.value],
            "E": arrayRow[EntryList.Col_PackType.value],
            "F": strRelLink,
        })

        summaryws.append({
            "A": arrayRow[EntryList.Col_Product.value],
            "B": arrayRow[EntryList.Col_DocName.value],
            "C": arrayRow[EntryList.Col_DocType.value],
            "D": arrayRow[EntryList.Col_PackType.value],
            "E": arrayRow[EntryList.Col_FilterComp.value],
            "F": '=HYPERLINK("./' + strRelLink + '")',
        })

        # wrap text on doc name cell and color the hyperlink cell
        summaryws.cell(row=summaryws.max_row, column=2).alignment = Alignment(wrap_text=True)
        summaryws.cell(row=summaryws.max_row, column=6).font = Font(
            bold=False,
            color=openpyXlColors.Color_DarkBlue.value,
        )

    # ------------------------------------------------------------------
    # Auto-fit column widths on both sheets (cap at 80 to avoid extremes)
    # (same dims pattern as Jochen, plus our width cap)
    # ------------------------------------------------------------------
    for ws in [selectionws, summaryws]:
        dims = {}
        for wsrow in ws.rows:
            for cell in wsrow:
                if cell.value:
                    dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
        for col, value in dims.items():
            ws.column_dimensions[col].width = min(value + 2, 80)

    wb.save(strSummaryPath)
    logger.info("I072 Summary Excel saved: \"" + strSummaryPath + "\"")
    return strSummaryPath


# =============================================================================
# STREAMLIT SESSION STATE HELPERS
# =============================================================================

def init_state():
    """Initialize session state keys if they don't exist yet."""
    if "selected_packtypes" not in st.session_state:
        st.session_state.selected_packtypes = set()
    if "selected_components" not in st.session_state:
        st.session_state.selected_components = set()


def clear_all():
    """Clear all user selections from session state."""
    st.session_state.selected_packtypes  = set()
    st.session_state.selected_components = set()


# =============================================================================
# MAIN STREAMLIT APP
# =============================================================================

def main():
    st.set_page_config(page_title="GEP Package Automation Tool", layout="wide")
    init_state()

    st.title("GEP Package Automation Tool")

    # ---- File upload ----
    uploaded_mdl = st.file_uploader("Upload MDL Excel file", type=["xlsx"])

    if uploaded_mdl is None:
        st.info("Please upload an MDL Excel file to proceed.")
        return

    # parse the uploaded MDL - returns plain lists, not dataclasses
    arrayEntries, setProducts, setPackageTypes = parse_mdl(uploaded_mdl)

    # ---- Sidebar: Review panel ----
    with st.sidebar:
        st.header("Review")

        intPtCount = len(st.session_state.selected_packtypes)
        with st.expander("Package type (" + str(intPtCount) + ")", expanded=False):
            if intPtCount:
                for strPt in sorted(st.session_state.selected_packtypes):
                    st.write("- " + strPt)
            else:
                st.caption("None selected (means ALL package types at build time).")

        listSelectedSorted = sorted(
            st.session_state.selected_components, key=lambda x: (x[0].lower(), x[1].lower())
        )
        intCompCount = len(listSelectedSorted)

        with st.expander("Components selected (" + str(intCompCount) + ")", expanded=False):
            if not listSelectedSorted:
                st.caption("No components selected yet.")
            else:
                st.caption("Click x to remove:")
                listToShow = listSelectedSorted[:SIDEBAR_COMPONENTS_RENDER]

                for strProduct, strComp in listToShow:
                    row = st.columns([12, 1])
                    with row[0]:
                        st.write(strProduct + " - " + strComp)
                    with row[1]:
                        if st.button("x", key="rm_" + strProduct + "_" + strComp, help="Remove"):
                            st.session_state.selected_components.discard((strProduct, strComp))
                            st.rerun()

                if intCompCount > SIDEBAR_COMPONENTS_RENDER:
                    st.caption("Showing first " + str(SIDEBAR_COMPONENTS_RENDER) + ".")

        st.markdown("---")
        if st.button("Clear all", use_container_width=True):
            clear_all()
            st.rerun()

    # ---- Main content area ----
    top    = st.container(border=True)
    mid    = st.container(border=True)
    bottom = st.container(border=True)

    # ---- Package Type selection ----
    with top:
        st.subheader("Package Type selection")

        listSelectedPacktypes = st.multiselect(
            "Package Type (Preliminary, Permitting or Operational)",
            options=sorted(setPackageTypes),
            default=sorted(st.session_state.selected_packtypes),
            help="Independent selection. Component options will NOT change based on this.",
        )
        st.session_state.selected_packtypes = set(listSelectedPacktypes)

    # ---- Component selection ----
    with mid:
        st.subheader("Add components")

        col1, col2 = st.columns([1, 2])

        with col1:
            strPickProduct = st.selectbox("Product Type", options=sorted(setProducts))

        with col2:
            listOptions = components_for_product(arrayEntries, strPickProduct)

            if not listOptions:
                st.warning("No components found for this sheet.")
            else:
                if len(listOptions) >= MAX_COMPONENT_OPTIONS_PER_SHEET:
                    st.caption(
                        "Showing first " + str(MAX_COMPONENT_OPTIONS_PER_SHEET) + " options (alphabetical)."
                    )

                with st.form("add_components_form", clear_on_submit=True):
                    listTempComponents = st.multiselect(
                        "Components in " + strPickProduct, options=listOptions
                    )
                    boolSubmitted = st.form_submit_button("Add selected components")

                    if boolSubmitted:
                        for strComp in listTempComponents:
                            st.session_state.selected_components.add((strPickProduct, strComp))
                        st.success("Added " + str(len(listTempComponents)) + " component(s).")
                        st.rerun()

    # ---- Build package ----
    with bottom:
        st.subheader("Build package")

        strRepoRoot    = st.text_input("Repository root", value="", placeholder="Path to your repository root folder")
        strOutputRoot  = st.text_input("Output folder (package destination)", value="", placeholder=r"Example: C:\Project\Package_Output")
        strProjectName = st.text_input("Project name", value="", placeholder="Example: MyGEPProject")

        if not FUZZY_AVAILABLE:
            st.caption(
                "Note: thefuzz library not installed - fuzzy match suggestions in the log will be disabled. "
                "Install with: pip install thefuzz"
            )

        if st.button("Build package", type="primary"):

            if not st.session_state.selected_components:
                st.warning("Please select at least one component to build the package.")
                return
            if not strRepoRoot or not os.path.isdir(strRepoRoot):
                st.warning("Please enter a valid repository root path.")
                return
            if not strOutputRoot:
                st.warning("Please enter a valid output folder path.")
                return
            if not strProjectName.strip():
                st.warning("Please enter a valid project name.")
                return

            os.makedirs(strOutputRoot, exist_ok=True)

            with st.spinner("Building package..."):
                try:
                    # step 1: scan the repository for all available files
                    arrayRepoFiles  = scan_repository(strRepoRoot)
                    intScannedCount = len(arrayRepoFiles)

                    # step 2: match each MDL file entry to a repository file
                    intMatchedCount, intNotFoundCount = match_mdl_files_with_repository(
                        arrayEntries, arrayRepoFiles
                    )

                    # step 3: copy matching files into the output folder
                    intCopied, listMissing, arrayCopiedEntries = build_package(
                        arrayEntries          = arrayEntries,
                        strOutputRoot         = strOutputRoot,
                        strProjectName        = strProjectName.strip(),
                        setSelectedPackTypes  = st.session_state.selected_packtypes,
                        setSelectedComponents = st.session_state.selected_components,
                    )

                    # step 4: create the summary Excel
                    strSummaryExcelPath = export_generated_summary_excel(
                        strOutputRoot      = strOutputRoot,
                        strProjectName     = strProjectName.strip(),
                        arrayCopiedEntries = arrayCopiedEntries,
                    )

                except ValueError as err:
                    logger.error("E090 ValueError during build: " + str(err))
                    st.error(str(err))
                    return

            st.success(
                "Package build complete! "
                + str(intCopied) + " file(s) copied, "
                + str(len(listMissing)) + " missing, "
                + str(intNotFoundCount) + " MDL file(s) not matched in repository, "
                + "out of " + str(intScannedCount) + " scanned."
            )
            st.info(
                "Package location: " + os.path.abspath(strOutputRoot)
                + "  |  Log file: " + strLoggingFile
            )

            if strSummaryExcelPath:
                st.success("Summary Excel created at: " + strSummaryExcelPath)
            else:
                st.info("No files were copied, so no summary Excel was created.")

            if listMissing:
                st.warning(str(len(listMissing)) + " MDL files were not found in the repository.")
                with st.expander("See missing files list"):
                    for strMissing in listMissing[:300]:
                        st.write(strMissing)
                    if len(listMissing) > 300:
                        st.caption("Showing first 300 of " + str(len(listMissing)) + " missing files.")


# =============================================================================
# LOGGER TEARDOWN
# Close and remove handlers cleanly on exit (same pattern as Jochen).
# =============================================================================

import atexit

def _close_logger():
    logger.info("*********************************************************************")
    logger.info("I099 GEP Package Automation Tool session finished.")
    logger.info("*********************************************************************")
    file_handler.close()
    stdout_handler.close()
    logger.removeHandler(file_handler)
    logger.removeHandler(stdout_handler)

atexit.register(_close_logger)


if __name__ == "__main__":
    main()
